'''
    @Author:  Yuxin Miao
    @Date:  2020-08-11 11:27:58
    @Last Modified by:  Yuxin Miao
    @Last Modified time:  2020-08-11 11:27:58
'''
import os
import xml.sax
import xml.etree.ElementTree as ET
import openpyxl

class DictHandler(xml.sax.ContentHandler):
    def __init__(self):
        self.word = None
        self.pronouce = []
        self.pronoucekk = []
        self.pos = None
        self.in_head = False
        self.has_read = False
    def startElement(self, tag, attributes):
        print(tag)
        self.CurrentData = tag
        if tag == 'Entry':
            print('******* new word *********')
            self.in_head = True
            self.has_read = False


    def endElement(self, tag):
        
        if not self.in_head:
            return

        if tag == 'Head':
            self.in_head = False
            self.has_read = False

        if tag == 'Entry':
            print('*******End********\n')

        if tag == 'PronCodes':
            self.has_read = True
            
        if tag == 'HWD':
            print('word:', self.word)

        elif tag == 'PRON' and not self.has_read:   
            print('pron:', self.pronouce)

        elif tag == 'PRONKK'and not self.has_read:
            print('pronkk', self.pronoucekk)
            if self.word == 'abominable':
                quit()
        
    def characters(self, content):
        if not self.in_head:
            return

        if self.CurrentData == 'HWD':
            self.word = content 
        elif self.CurrentData == 'PRON' and not self.has_read:   
            self.pronouce.append(content)
        elif self.CurrentData == 'PRONKK'and not self.has_read:
            self.pronoucekk.append(content)
        # elif self.CurrentData == 'POS':
        #     self.pos = content



def ETparser(root):
    word = ''
    pron = []
    pronkk = []
    temp = []

    for childroot in root:
        if childroot.tag == 'Head':
            for childhead in childroot:
                if childhead.tag == 'HWD':
                    word = childhead.text
                if childhead.tag == 'PronCodes':
                    for childpron in childhead:
                        if childpron.tag != 'PRON' and childpron.tag!= 'PRONKK':
                            continue 
                        temp = []
                        try:
                            temptext = childpron.text.replace(' ','').replace("\n","").replace("\t","")
                            temp.append(temptext)
                        except:
                            # print('head',word)
                            pass
                        for childi in childpron:
                            # print(childi.text)
                            temp.append(childi.text.replace(' ','').replace("\n","").replace("\t",""))
                            try:
                                temp.append(childi.tail.replace(' ','').replace("\n","").replace("\t",""))
                            except:
                                # print('tail: ',word)
                                pass
                        temppron = ''.join(temp)

                        if childpron.tag == 'PRON':
                            pron.append(temppron)
                        else:
                            pronkk.append(temppron)
                
    word_list.append(word)
    pron_list.append(pron)
    pronkk_list.append(pronkk)


if __name__ == "__main__":
    # parser = xml.sax.make_parser()

    # parser.setFeature(xml.sax.handler.feature_namespaces, 0)

    # handler = DictHandler()
    # parser.setContentHandler(handler)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'LongMan'

    word_list = []
    pron_list = []
    pronkk_list = []

    files = os.listdir('data/raw/')
    files.sort()

    
    for f in files:
    # parser.parse('data/raw/test.xml')
    # # parser.parse('data/raw/'+files[0])
        name = 'data/raw/'+f
        print(name)
        parser = ET.XMLParser(encoding="utf-8")
        tree = ET.parse(name,parser=parser)
        root = tree.getroot()
        for child in root:
            ETparser(child)

    # print(word_list)
    # print(pron_list)
    # print(pronkk_list)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'LongMan'
    title = ['word','pronounce','pronouncekk']
    for i in range(len(title)):
        sheet.cell(1, i+1,title[i])
            
    for i in range(len(word_list)):
        sheet.cell(i+2,1,word_list[i])

    for i in range(len(pron_list)):
        sheet.cell(i+2,2,','.join(pron_list[i]))
            
    for i in range(len(pronkk_list)):
        sheet.cell(i+2,3,','.join(pronkk_list[i]))
    
    workbook.save('longman.xlsx')
    print('ooooooooooooooooooooooooooooooooooooooooooooooooh!')
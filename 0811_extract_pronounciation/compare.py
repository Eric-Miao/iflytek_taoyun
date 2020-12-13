'''
    @Author:  Yuxin Miao
    @Date:  2020-08-11 17:20:04
    @Last Modified by:  Yuxin Miao
    @Last Modified time:  2020-08-11 17:20:04
'''
import openpyxl

INPUT = 'longman_extraction.xlsx'
SOURCE = 'input.xlsx'
OUTPUT = 'output.xlsx'

def addtwodimdict(thedict, key_a, key_b, val): 
    if key_a in thedict:
        thedict[key_a].update({key_b: val})
    else:
        thedict.update({key_a:{key_b: val}})

if __name__ == "__main__":

    dic = openpyxl.load_workbook(INPUT)
    sheet_dict = dic.active
    source = openpyxl.load_workbook(SOURCE)
    sheet_pen = source.active
    
    longman = dict()
    word_db = list(sheet_dict.columns)[0]
    pron_db = list(sheet_dict.columns)[1]
    pronkk_db = list(sheet_dict.columns)[2]

    for i in range(len(word_db)):
        # try:
        #     assert(type(word_db[i].value)=='str')
        # except:
        #     print('error in dict')
        #     print(word_db[i].value)
        #     print(type(word_db[i].value))
        addtwodimdict(longman,word_db[i].value,'pron',pron_db[i].value)
        addtwodimdict(longman,word_db[i].value,'pronkk',pronkk_db[i].value)
    
    word_list = list(sheet_pen.columns)[1]
    
    print(longman)
    assert(isinstance(longman,dict))
    assert(isinstance(longman[0],dict))
    quit()

    for word in word_list:
        if word.value in longman:
            # try:
            #     temp = word_list[word.value]
            # except:
            #     print('error in refer')
            #     print(word.value)
            #     print(type(word.value))
            #     continue
            if word_list[word.value]['pron']:
                sheet_pen.cell(word.row,word.column+1,word_list[word.value]['pron'])
            if word_list[word.value]['pronkk']:
                sheet_pen.cell(word.row,word.column+2,word_list[word.value]['pronkk'])
            sheet_pen.cell(word.row,word.column+3,'modified')
    source.save(OUTPUT)                                                                                                                                                                             


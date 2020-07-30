# -*- coding:utf-8 -*-  

'''
    @Author:  Yuxin Miao
    @Date:  2020-07-28 09:21:37
    @Last Modified by:  Yuxin Miao
    @Last Modified time:  2020-07-28 09:21:37
'''
import openpyxl
import json

DATA = 'sentence.xlsx'


if __name__ == "__main__":
    data = openpyxl.load_workbook(DATA)
    sheet = data.worksheets[0]

    example = list(sheet.columns)[2]

    assert(example[0].value == 'example')
    for cell in example:

        raw = cell.value
        if raw=='example':
            continue
        processed_txt = raw
        processed_txt = processed_txt.replace('[','')
        processed_txt = processed_txt.replace(']','')
        processed_txt = processed_txt.replace('},{','}|{')
        processed_txt = processed_txt.replace('\n','')
        processed_txt = processed_txt.split('|')
        


        new_json_list=[]
        new_json=None
        error_flag=0
        # print(processed_txt)
        for each in processed_txt:
            try:
                case = json.loads(each)
            except:
                print(cell, "error exists")
                error_flag=1
                continue



            if 'orgin' in case:
                case["origin"] = case.pop("orgin")

            if not 'origin' in case or case['origin']==None:
                if not 'orginDetail' in case or case['orginDetail'] == None:
                    continue
                try:
                    od = case['orginDetail']
                    case['origin'] = od.split('-')[0]
                except:
                    print(cell, "None type error exists\n")
                    continue

            case['origin'] = case['origin'].replace('文本','')
            case['orginDetail'] = case['orginDetail'].replace('文本','')

            case["orgin"] = case.pop("origin")

            new_json_list.append(json.dumps(case,ensure_ascii=False))
        
        new_json = '[' + ','.join(new_json_list) + ']'

        if not error_flag:
            sheet.cell(cell.row,cell.column,new_json)
        else:
            print(cell,'naive replace\n')
            sheet.cell(cell.row,cell.column,cell.value.replace('文本',''))
    
    data.save(filename="newData.xlsx")

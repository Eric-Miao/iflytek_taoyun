'''
    @Author:  Yuxin Miao
    @Date:  2020-07-31 17:00:42
    @Last Modified by:  Yuxin Miao
    @Last Modified time:  2020-07-31 17:00:42
'''

'''
❶❷❸❹❺❻❼❽❾❿
āáǎàōóǒòēéěèīíǐìūúǔùüǖǘǚǜêê̄ếê̌ề
ĀÁǍÀŌÓǑÒĒÉĚÈĪÍǏÌŪÚǓÙÜǕǗǙǛÊÊ̄ẾÊ̌Ề

re.findall(r'[\u4e00-\u9fa5]{1,}\s*(.*)',test)

汉字表达式：[\u4e00-\u9fa5]
拼音表达式：[Aa-zZāáǎàōóǒòēéěèīíǐìūúǔùüǖǘǚǜńňǹḿmg]*
字符表达式：[a-zA-Z0-9_]
包含中英文标点符号和其他特殊符号的表达式：[\W]
'''

import json
import openpyxl
import re
from bloom_filter import BloomFilter as BF

DB = 'iflytek_db.xlsx'
DICT = 'XDHYCD7th.txt'
BIGDICT = 'DCD.txt'
OUTPUT = 'output.txt'
RESULT = 'result.xlsx' 
LOG = 'log.txt'

error_explanation = []
error_pinyin = []

word_list =[]
pinyin_list = []
mean_list=[]
stupid_symbol = ['\n','（～儿）','∥','•','’','∥','（～的）','▲','*','⁰','¹','²','³','⁴','⁵','⁶','⁷','⁸','⁹','₀','₁','₂','₃','₄','₅','₆','₇','₈','₉']

def no_explanation():
    pass
def no_pinyin():
    pass
def DCD_get_pinyin(line):
    return re.findall(r'[\u4e00-\u9fa5]{1,}[ \t]*(.*)',line)[0]

def DCD_get_word(line):
    return re.findall(r'[\u4e00-\u9fa5]{1,}',line)[0]

def extract_explanation(text):
    # print(text)
    for i in stupid_symbol:
        text=text.replace(i,'')
    # print(text)
    text = re.sub(r'[（\(].*?[\)）]','',text)
    # print(text)
    text = re.sub(r'[＜〈].*?[〉＞]','',text)
    # print(text)
    text = re.sub(r'[另]?见[0-9]*.*?。$','',text)
    # print(text)
    text = re.sub(r'参看[0-9]*.*?。$','',text)
    # print(text)
    line=re.sub(r'【.*?】[ \t]?.*?[Aa-zZ)āáǎàōóǒòēéěèīíǐìānkānɡɡūúǔùüǖǘǚǜńňǹḿmêê̄ếê̌ềĀÁǍÀŌÓǑÒĒÉĚÈĪÍǏÌŪÚǓÙÜǕgǗǙǛÊÊ̄ẾÊ̌Ềɑnɑo❶…āijiěānkānběnběnbēishìBāuàjiàoBáiliánjiàbāihuɑiāijiěbáqiābāihuɑibàntóu]*','',text)
    if line == '':
        return []
    try:
        # temp1=re.split('❶|❷|❸|❹|❺|❻|❼|❽|❾|❿',line)
        temp1=re.split(r'[❶❷❸❹❺❻❼❽❾❿]',line)
        # print('allright')
        # print(temp1)
        while '' in temp1:
            temp1.remove('')
    except:
        print('error',line)
        quit()

    while '' in temp1:
        temp1.remove('')
    while '\n' in temp1:
        temp1.remove('\n')
        
    
    return temp1
    

def extraction():
    iflytek_db = openpyxl.load_workbook(DB)
    
    log_file=open(LOG,'w')

    db = iflytek_db.worksheets[0]
    hanyu = open(DICT,'r')
    hanyu_big = open(BIGDICT,'r')
    out = open(OUTPUT,'w')

    hanyudict = BF(210000,0.01)
    hanyubigdict = {}
    
    result_book = openpyxl.Workbook()
    sheet_available = result_book.create_sheet()
    sheet_available.title = 'available'
    sheet_no_pinyin = result_book.create_sheet()
    sheet_no_pinyin.title = 'missing pinyin'
    sheet_no_explanation = result_book.active
    sheet_no_explanation.title = 'missing explanation'

    
    err=0
    for line in hanyu_big.readlines():
        if '【' in line:
            continue
        try:
            word = DCD_get_word(line)
            if len(word) == 1:
                continue
            hanyubigdict[DCD_get_word(line)] = DCD_get_pinyin(line)
            # print(re.findall(r'[\u4e00-\u9fa5]{1,}',line))
            # print(re.findall(r'[\u4e00-\u9fa5]{1,}[ \t]*(.*)',line))
            # quit()
        except:
            pass


    word_db = list(db.columns)[1]
    assert(word_db[0].value == 'word')

    for cell in word_db:
        if cell.value == 'word':
            continue
        hanyudict.add(cell.value)

    # Start to compare
    cnt_missing = 0
    cnt_total = 0
    cnt_single = 0
    cnt_no_pinyin = 0
    cnt_good = 0
    cnt_no_exp = 0
    for line in hanyu.readlines():
        if '】' not in line:
            continue
        if '现用替代字' in line:
            continue
        # try:
        #     word = re.findall(r'【(.*?)】',line)[0]
        # except:
        #     print('**************\nerror')
        #     print(line)
        #     print(word)
        #     quit(0)
        word = re.findall(r'【(.*?)】',line)[0]
        if len(word) == 1:
            cnt_single = cnt_single + 1
            continue            

        else:
            if not word in hanyudict:
                cnt_missing += 1
                try:
                    res=extract_explanation(line)
                except:
                    print('error extraction ',line)
                    quit()
                if len(res) == 0:
                    error_explanation.append(word)
                    cnt_no_exp  += 1
                else:
                    temp = {"partOfSpeech":'',"mean":"","example":"","word_explain":""}
                    for each in res:
                        temp['mean'] = each
                    mean_list.append('[' + json.dumps(temp, separators=(',', ':'), ensure_ascii=False) + ']')
                    word_list.append(word)
                    cnt_total += 1
                    if not word in hanyubigdict:
                        # no pinyin take down
                        pinyin_list.append('')
                        error_pinyin.append(word)
                        cnt_no_pinyin += 1
                    else:
                        pinyin_list.append(hanyubigdict[word])
                        cnt_good += 1


    title = ['word','pinyin拼音','synonym同义词','antonym反义词','basicmean释义','detailMean','classification',\
        'word_type','origin','example','是否检查']

    for i in range(len(title)):
        sheet_available.cell(1, i+1,title[i])
    
    for i in range(len(word_list)):
        sheet_available.cell(i+2,1,word_list[i])

    for i in range(len(pinyin_list)):
        sheet_available.cell(i+2,2,pinyin_list[i])
    
    for i in range(len(mean_list)):
        sheet_available.cell(i+2,5,mean_list[i])



    title = ['id','word']
    for i in range(len(title)):
        sheet_no_explanation.cell(1, i+1,title[i])
    for i in range(len(error_explanation)):
        sheet_no_explanation.cell(i+2,1,i)
    for i in range(len(error_explanation)):
        sheet_no_explanation.cell(i+2,2,error_explanation[i])
     


    for i in range(len(title)):
        sheet_no_pinyin.cell(1, i+1,title[i])
    for i in range(len(error_pinyin)):
        sheet_no_pinyin.cell(i+2,1,i)
    for i in range(len(error_pinyin)):
        sheet_no_pinyin.cell(i+2,2,error_pinyin[i])
    # When find one that doesn't exist in database, go check in DCD for pinyin, go check in hanyudict for explanation.
    # Before check for pinyin, check for existence first then search for pinyin to save time avoid no help search.
    # Only split and process the explanation after confirming that one is missing.
    out.close()
    log_file.close()
    result_book.save(RESULT)
    print('Total missing: %d \nTotal adding: %d \n  With Good adding: %d \n   With no Pinyin:%d \nNo exp: %d \nSingle Character: %d' % (cnt_missing,cnt_total,cnt_good,cnt_no_pinyin,cnt_no_exp,cnt_single))


if __name__ == "__main__":
    # test = '【大锅饭】dàɡuōfàn〈名〉❶供多数人吃的普通伙食。❷见179页【吃大锅饭】。'
    # extract_explanation(test)

    extraction()
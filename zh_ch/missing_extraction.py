# -*- coding:utf-8 -*-  

'''
    @Author:  Yuxin Miao
    @Date:  2020-07-29 18:28:09
    @Last Modified by:  Yuxin Miao
    @Last Modified time:  2020-07-29 18:28:09
'''
import openpyxl
import json

DATA = 'input.xlsx'
OUTPUT = 'output.xlsx'
TARGET = ['prefix','other','suffix','idiom']

if __name__ == '__main__':
    output = openpyxl.Workbook()
    sheet_raw = output.active
    sheet_raw.title = 'raw'
    sheet_prefix = output.create_sheet()
    sheet_prefix.title = 'missing prefix'
    sheet_other = output.create_sheet()
    sheet_other.title = 'missing other'
    sheet_suffix = output.create_sheet()
    sheet_suffix.title = 'missing suffix'
    sheet_idiom = output.create_sheet()
    sheet_idiom.title = 'missing idiom'
    sheet_all = output.create_sheet()
    sheet_all.title = 'missing all'


    output_sheets = {'prefix':sheet_prefix,'suffix':sheet_suffix,\
        'other':sheet_other,'idiom':sheet_idiom,}

    data = openpyxl.load_workbook(DATA)
    sheet = data.worksheets[0]

    # char_list = list(sheet.columns)[1]
    # pinyin_list = list(sheet.columns)[2]
    phrase_list = list(sheet.columns)[3]

    for target in TARGET:
        output_sheet = output_sheets[target]
        row=1
        for each in phrase_list:
            if each.value == 'relatedTerm':
                continue
            try:
                case = json.loads(each.value)
            except:
                continue

            if not case[target] or not target in case:
                output_sheet.cell(row=row,column=1,value=sheet.cell(each.row,1).value)
                output_sheet.cell(row=row,column=2,value=sheet.cell(each.row,2).value)
                output_sheet.cell(row=row,column=3,value=sheet.cell(each.row,3).value)
                row = row + 1


    row=1
    for each in phrase_list:
        if each.value == 'relatedTerm':
            continue
        try:
            case = json.loads(each.value)
        except:
            sheet_all.cell(row=row,column=1,value=sheet.cell(each.row,1).value)
            sheet_all.cell(row=row,column=2,value=sheet.cell(each.row,2).value)
            sheet_all.cell(row=row,column=3,value=sheet.cell(each.row,3).value)
            row = row + 1
            continue

        if case[TARGET[0]] or case[TARGET[1]] or case[TARGET[2]] or case[TARGET[3]]:
            pass
        else:
            sheet_all.cell(row=row,column=1,value=sheet.cell(each.row,1).value)
            sheet_all.cell(row=row,column=2,value=sheet.cell(each.row,2).value)
            sheet_all.cell(row=row,column=3,value=sheet.cell(each.row,3).value)
            
            row = row + 1

    output.save(filename=OUTPUT)
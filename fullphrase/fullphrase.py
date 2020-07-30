'''
    @Author:  Yuxin Miao
    @Date:  2020-07-27 18:18:24
    @Last Modified by:  Yuxin Miao
    @Last Modified time:  2020-07-27 18:18:24
'''

import openpyxl

DATA = 'phrase.xlsx'


if __name__ == "__main__":
    data = openpyxl.load_workbook(DATA)
    sheet = data.worksheets[0]
    ph=None
    fullph=None
    fullphpy=None
    fullphtts=None
    col_ph=0
    col_fullph=0
    col_fullphpy=0
    col_fullphtts=0
    
    
    for col in sheet.columns:
        cell = col[0]
        column = cell.column

        if cell.value == 'word':
            ph=col
            col_ph = column

        if cell.value == 'fullPhrase':
            fullph = col
            col_fullph = column

        elif cell.value == 'fullPhrasePinyin':
            fullphpy = col
            col_fullphpy = column

        elif cell.value == 'fullPhraseTts':
            fullphtts = col
            col_fullphtts = column

    fullrow = 0
    halfrow = 0
    for cell in ph:
        word = cell.value
        if  '，' in word:
            fullrow=cell.row
            temp = word.split('，')

            for half in temp:
                for sec_cell in ph:
                    if sec_cell.value == half:
                        halfrow = sec_cell.row
                        if not sheet.cell(row=halfrow,column=col_fullph).value:
                            sheet.cell(row=halfrow,column=col_fullph,value=sheet.cell(fullrow,col_ph).value)         
                            sheet.cell(row=halfrow,column=col_fullphpy,value=sheet.cell(fullrow,col_ph+1).value)  
                        sheet.cell(row=halfrow,column=col_fullphtts,value=sheet.cell(fullrow,col_ph).value)               

    
    data.save(filename="newData.xlsx")
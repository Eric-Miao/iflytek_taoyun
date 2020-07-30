'''
    @Author:  Yuxin Miao
    @Date:  2020-07-30 08:50:54
    @Last Modified by:  Yuxin Miao
    @Last Modified time:  2020-07-30 08:50:54
'''

import openpyxl
import json


INPUT = 'input.xlsx'
OUTPUT = 'output_shorten.xlsx'
LOG = 'deleted.xlsx'

if __name__ == "__main__":
    data = openpyxl.load_workbook(INPUT)
    sheet = data.worksheets[0]
    sheet.title = 'shorten senteces'

    log = openpyxl.Workbook()
    deleted_sheet = log.active
    deleted_sheet.title = 'deleted sentences'

    log_list = []
    num_list = []
    char_list = []

    example = list(sheet.columns)[3]
    assert(example[0].value == 'Example')

    for cell in example:
        if not cell.value or cell.value == 'Example':
            if not cell.value == 'Example':
                print(cell,'no content')
            continue

        raw = cell.value

        # pre-process of the content
        processed_txt = raw
        processed_txt = processed_txt.replace('[','')
        processed_txt = processed_txt.replace(']','')
        processed_txt = processed_txt.replace('},','}|')
        processed_txt = processed_txt.replace('\n','')
        processed_txt = processed_txt.split('|')
        while '' in processed_txt:
            processed_txt.remove('')
        new_json_list=[]
        new_json=None

        
        for each in processed_txt:
            try:
                case = json.loads(each)
            except:
                print(processed_txt)
                print(each)
                print(cell, "json.loads error exists")
                quit() 

            if len(case['key']) < 50:
                new_json_list.append(json.dumps(case,ensure_ascii=False))

            else:
                log_list.append(case['key'])
                num_list.append(sheet.cell(cell.row,1).value)
                char_list.append(sheet.cell(cell.row,2).value)
                continue
        
        if len(new_json_list) == 0:
            for i in range(len(processed_txt)):
                log_list.pop()
                num_list.pop()
                char_list.pop()
            continue

        new_json = '[' + ','.join(new_json_list) + ']'
        sheet.cell(cell.row,cell.column,new_json)



    title=['Id','Word', 'Example']

    for i in range(len(title)):
        deleted_sheet.cell(1, i+1,title[i])
    
    for i in range(len(num_list)):
        deleted_sheet.cell(i+2,1,num_list[i])

    for i in range(len(char_list)):
        deleted_sheet.cell(i+2,2,char_list[i])
    
    for i in range(len(log_list)):
        deleted_sheet.cell(i+2,3,log_list[i])


    data.save(filename=OUTPUT)
    log.save(LOG)
    
    # sanity check
    
    print('********************\nstart Sanity Check\n')
    data = openpyxl.load_workbook(OUTPUT)
    sheet = data.worksheets[0]

    log = openpyxl.load_workbook(LOG)
    not_deleted_sheet = log.create_sheet()
    not_deleted_sheet.title = 'over 50'


    example = list(sheet.columns)[3]

    assert(example[0].value == 'Example')

    log_list = []
    num_list = []
    char_list = []
    
    for cell in example:
        if not cell.value or cell.value == 'Example':
            if not cell.value == 'Example':
                print(cell,'no content')
            continue

        raw = cell.value

        # pre-process of the content
        processed_txt = raw
        processed_txt = processed_txt.replace('[','')
        processed_txt = processed_txt.replace(']','')
        processed_txt = processed_txt.replace('},','}|')
        processed_txt = processed_txt.replace('\n','')
        processed_txt = processed_txt.split('|')
        while '' in processed_txt:
            processed_txt.remove('')
        if len(processed_txt) == 0:
            print('there are cases of no examples', cell)
            quit()
        error_flag=0
        for each in processed_txt:
            try:
                case = json.loads(each)
            except:
                print(processed_txt)
                print(each)
                print(cell, "json.loads error exists")
                quit() 
            # print(case['key'])
            if len(case['key']) < 50:
                continue
            else:
                error_flag = error_flag + 1
                # print('over 50')
                log_list.append(case['key'])
                num_list.append(sheet.cell(cell.row,1).value)
                char_list.append(sheet.cell(cell.row,2).value)

        if error_flag == 0:
            continue
        elif error_flag == len(processed_txt):
            # if error_flag > 1:
                # print(cell,'has all long examples')
                # print(cell.value)
                # quit()
            continue
        else:
            print('error at ', cell)
            quit()

    title=['Id','Word', 'Example']

    for i in range(len(title)):
        not_deleted_sheet.cell(1, i+1,title[i])
    
    for i in range(len(num_list)):
        not_deleted_sheet.cell(i+2,1,num_list[i])

    for i in range(len(char_list)):
        not_deleted_sheet.cell(i+2,2,char_list[i])
    
    for i in range(len(log_list)):
        not_deleted_sheet.cell(i+2,3,log_list[i])

    log.save(LOG)